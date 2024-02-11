from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get('https://iotvega.com/product')

product = driver.find_element(By.CLASS_NAME, 'main-container')
product.click()

name_product = driver.find_element(By.XPATH, "/html/body/section[1]/div/div/div/div/div/h1").text

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1).value = "Характеристики"
sheet.cell(row=1, column=2).value = name_product

i = 2
while i < 42:
    for N in range(1, 40, 2):
        caracteristic = driver.find_element(By.XPATH, f'/html/body/section[3]/div/div/div[1]/table/tbody/tr[{N}]/td[1]').text
        value = driver.find_element(By.XPATH, f'/html/body/section[3]/div/div/div[1]/table/tbody/tr[{N}]/td[2]').text

        sheet.cell(row=i, column=1).value = caracteristic
        sheet.cell(row=i, column=2).value = value

        i += 1

workbook.save("res.xlsx")

time.sleep(5)
driver.quit()